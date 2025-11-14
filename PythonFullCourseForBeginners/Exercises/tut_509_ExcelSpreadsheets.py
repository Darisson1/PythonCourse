import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Tabelle1']
    # its possible to give the coordinates as A1
    cell = sheet['A1']
    # or like this:
    cell = sheet.cell(1,1)
    print(cell.value)

    # how to get the number of rows
    # print(sheet.max_row)

    # loop throgh the rows but begin with 2,
    # because we dont want the title, only the values
    for row in range(2, sheet.max_row + 1):
        # we get the row from the loop and the column with the prices is 3
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.number_format = '#,##0.00 [$€-407]'
        corrected_price_cell.value = corrected_price

    values = Reference( sheet,
                        min_row=2,
                        max_row=sheet.max_row,
                        min_col=4,
                        max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save(filename)


process_workbook('transactions.xlsx')
